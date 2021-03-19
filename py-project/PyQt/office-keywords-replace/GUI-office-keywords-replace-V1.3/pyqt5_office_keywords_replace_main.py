#导入程序运行必须模块
from win32com.client import Dispatch
import sys
import os
import time
import shutil
import openpyxl as xl
import xlrd
from xlutils.copy import copy


from PyQt5 import QtCore, QtGui
from PyQt5.QtCore import QEventLoop, QTimer
from PyQt5.QtGui import QPainter, QColor
#PyQt5中使用的基本控件都在PyQt5.QtWidgets模块中
from PyQt5.QtWidgets import QWidget, QApplication, QMainWindow,QGroupBox, QPushButton, QLabel, QHBoxLayout, QVBoxLayout, QGridLayout, QFormLayout, QLineEdit, QTextEdit, QInputDialog, QFileDialog, QMessageBox, QDesktopWidget

#导入UI文件类
from pyqt5_office_keywords_replace_ui import Ui_MainWindow

class MyMainForm(QMainWindow, Ui_MainWindow):

    def __init__(self, parent=None):
        super(MyMainForm, self).__init__(parent)
        self.setupUi(self)

        # 添加选择替换目录 按钮信号和槽
        self.ori_dir_btn.clicked.connect(self.replacePath)

        # 添加添加选择存储目录 按钮信号和槽
        self.replace_dir_btn.clicked.connect(self.savePath)

        #添加开始转换按钮信号和槽
        self.start_transfer.clicked.connect(self.transfer)

        self.word = Dispatch('kwps.Application')


    #需要替换的目录
    def replacePath(self):
        path = QFileDialog.getExistingDirectory(self, "请选择您要替换的目录")
        # 判断选择的文件是否存在
        if os.path.exists(path):
            # 将保存url放入路径文本框中
            self.ori_dir_input.setText(path)
        else:
            self.showMsg('错误', '您选择的目录不存在，请重新选择！')
            return False

    def savePath(self):
        path = QFileDialog.getExistingDirectory(self, "请选择您要保存的位置")
        if os.path.exists(path):
            # 将保存url放入路径文本框中
            self.replace_dir_input.setText(path)
        else:
            self.showMsg('错误', '您选择的目录不存在，请重新选择！')
            return False

    # 显示消息
    def showMsg(self, title, content, icon=3):
        box = QMessageBox(QMessageBox.Warning, title, content)
        # 设置左上角消息框图标
        # box.setWindowIcon(QIcon(r'E:\site\python\cutimg\favicon.ico'))
        # 添加按钮，可用中文
        yes = box.addButton('确定', QMessageBox.YesRole)
        # 设置消息框中内容前面的图标
        box.setIcon(icon)
        # 显示该问答框
        box.exec()
        return False

    #开始转换
    def transfer(self):

        #需要替换的目录
        self.replace_dir = self.ori_dir_input.text()
        #保存的目录
        self.save_dir = self.replace_dir_input.text()
        #关键词，并且去除收尾空格
        self.keywords = self.replace_words.toPlainText().strip()

        if not os.path.isdir(self.replace_dir) or not os.path.isdir(self.save_dir):
            self.showMsg('错误', '您选择的目录不存在，请重新选择！')
            return False


        #存储目录必须为空目录
        if os.listdir(self.save_dir):
            self.showMsg('错误', '存储目录必须为空目录！')
            return False

        #文件目录和子目录不能相同
        if self.replace_dir == self.save_dir:
            self.showMsg('错误', '替换目录和存储目录不能相同！')
            return False

        # 替换目录和文件目录不能互为子目录
        if self.replace_dir.startswith(self.save_dir) or self.save_dir.startswith(self.replace_dir):
            self.showMsg('错误', '替换目录和存储目录不能相互包含！')
            return False

        if self.keywords == "":
            self.showMsg('错误', '关键词不能为空')
            return False

        #存储目录不存在就直接创建
        if not os.path.exists(self.save_dir):
            os.makedirs(self.save_dir)

        #开始替换
        self.replace_tool()
        #将转换结果输出到界面上
        # self.result_display.setPlainText(self.keywords)


    #删除目录及目录下的所有文件
    def del_dir(self, filepath):
        del_list = os.listdir(filepath)
        for f in del_list:
            file_path = os.path.join(filepath, f)
            if os.path.isfile(file_path):
                os.remove(file_path)
            elif os.path.isdir(file_path):
                shutil.rmtree(file_path)

    #将doc替换成docx
    def doc2docx(self, file):
        #获取文件目录，文件名带后缀，文件名，后缀
        file_dir, tmpfilename = os.path.split(file)
        file_name, extension = os.path.splitext(tmpfilename)
        doc = self.word.Documents.Open(file)


        new_file = file_dir + "/" +file_name + '.docx'
        doc.SaveAs(new_file, 12)  #12表示docx格式
        doc.Close()
        #删除doc文件
        os.remove(file)

        #docx文件转换
        self.docx_keywords_replace(new_file)

    #将目录及子目录下的文件copy到另一个目录
    def deep_copy_dir(self, origin_dir, target_dir):
        for files in os.listdir(origin_dir):
            name = os.path.join(origin_dir, files)
            back_name = os.path.join(target_dir, files)
            if os.path.isfile(name):
                if os.path.isfile(back_name):
                    shutil.copy(name, back_name)
                else:
                    shutil.copy(name, back_name)
            else:
                if not os.path.isdir(back_name):
                    os.makedirs(back_name)
                self.deep_copy_dir(name, back_name)

    #获取目录及子目录下的文件，并转换成可以处理的格式
    def recurse_transfer_file(self, path):
        for root, dirs, files in os.walk(path):
            for file in files:
                file = file.replace('\\', '/')
                src_file = os.path.join(root, file)

                #doc文件，先进行转换
                if src_file.endswith(".doc") and not src_file.startswith('~$'):
                    self.display_result("{src_file}格式转换".format(src_file=src_file))
                    self.doc2docx(src_file)
                elif src_file.endswith(".docx") and not src_file.startswith('~$'):
                    self.docx_keywords_replace(src_file)
                elif src_file.endswith(".xls") and not src_file.startswith('~$'):
                    # self.xls_keywords_replace(src_file)
                    pass
                elif src_file.endswith(".xlsx") and not src_file.startswith('~$'):
                    # self.xlsx_keywords_replace(src_file)
                    pass
                elif src_file.endswith(".pdf") and not src_file.startswith('~$'):
                    pass

    #word文件关键字替换
    def docx_keywords_replace(self, file):

        # 获取文件目录，文件名带后缀，文件名，后缀
        file_dir, tmpfilename = os.path.split(file)
        file_name, extension = os.path.splitext(tmpfilename)
        origin_file_name = file_name
        doc = self.word.Documents.Open(file)
        a = self.word.ActiveDocument.Sections

        # 每行
        every_line = self.keywords.split('\n')

        for i in range(len(a)):

            for line in every_line:
                split_list = line.split('|')
                old_word = split_list[0].strip().replace('\r', '').replace('\n', '').replace('\t', '')
                new_word = split_list[1].strip().replace('\r', '').replace('\n', '').replace('\t', '')

                #对页眉进行替换
                self.word.ActiveDocument.Sections[i].Headers[0].Range.Find.Execute(old_word, False, False, False, False, False, True, 1, True, new_word, 2)

                #对页脚进行替换
                self.word.ActiveDocument.Sections[i].Footers[0].Range.Find.Execute(old_word, False, False, False, False, False, True, 1, True, new_word, 2)

                #替换正文
                self.word.Selection.Find.Execute(old_word, False, False, False, False, False, True, 1, True, new_word, 2)

                file_name = file_name.replace(old_word, new_word)
                display_content = "{file}进行{old_word}->{new_word}替换".format(file=file, old_word = old_word, new_word = new_word)
                self.display_result(display_content)

        #存储替换文件
        doc.SaveAs(r"{0}/{1}.docx".format(file_dir, file_name))
        if origin_file_name != file_name:
            os.remove(file)
        doc.Close()
        # self.word.Quit()


    #替换xlsx关键词
    def xlsx_keywords_replace(self, file):
        wb = xl.load_workbook(file)
        ws = wb.worksheets[0]

        every_line = self.keywords.split('\n')
        file_dir, tmpfilename = os.path.split(file)
        file_name, extension = os.path.splitext(tmpfilename)

        # 每个关键词替换
        for line in every_line:

            split_list = line.split('|')
            old_word = split_list[0].strip().replace('\r', '').replace('\n', '').replace('\t', '')
            new_word = split_list[1].strip().replace('\r', '').replace('\n', '').replace('\t', '')

            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    content = ws.cell(row=row, column=col).value
                    ws.cell(row=row, column=col).value = content.replace(old_word, new_word, 1)

            file_name = file_name.replace(old_word, new_word)

        #存储
        wb.save(r"{0}/{1}{2}".format(file_dir, file_name,extension))
        #删除源文件
        os.remove(file)
        wb.close()
        display_content = file + "关键词转换成功"
        self.display_result(display_content)


    #替换xls文件关键词
    def xls_keywords_replace(self, file):

        wb = xlrd.open_workbook(file, formatting_info=True)  # 获取xls，保留原格式
        ws = wb.sheet_by_index(0)  # 根据index获取sheet
        rows = ws.nrows
        cols = ws.ncols
        newbook = copy(wb)  # 复制xls
        newsheet = newbook.get_sheet(0)

        every_line = self.keywords.split('\n')
        file_dir, tmpfilename = os.path.split(file)
        file_name, extension = os.path.splitext(tmpfilename)
        # 每个关键词替换
        for line in every_line:
            split_list = line.split('|')
            old_word = split_list[0].strip().replace('\r', '').replace('\n', '').replace('\t', '')
            new_word = split_list[1].strip().replace('\r', '').replace('\n', '').replace('\t', '')

            # 遍历每个单元格，进行替换操作
            for row in range(1, rows):
                for col in range(1, cols):
                    content = ws.cell(row, col).value
                    if (content != None and isinstance(content, str)):  # 判断不为空且为字符
                        if (content.find(old_word) != -1):  # 找到需要替换的字符
                            newsheet.write(row, col, content.replace(old_word, new_word))

            file_name = file_name.replace(old_word, new_word)

        # 保存新的xls以替换原有的xls
        newbook.save(r"{0}/{1}{2}".format(file_dir, file_name, extension))
        # 删除源文件
        os.remove(file)
        display_content = file + "关键词转换成功"
        self.display_result(display_content)

    #将处理结果实时显示到控制台
    def display_result(self,msg):
        self.result_display.append("{msg}\n".format(msg=msg))
        # self.cursor = self.result_display.textCursor()
        # self.result_display.moveCursor(self.cursor.End)
        QApplication.processEvents()

    #核心替换功能
    def replace_tool(self):
        try:
            self.result_display.clear()
            self.display_result("正进行目录清理")
            # 转换前先清空目录下文件
            self.del_dir(self.save_dir)
            self.display_result("正进行目录文件拷贝")
            #将原始目录及子目录文件全部拷贝至新目录
            self.deep_copy_dir(self.replace_dir, self.save_dir)

            self.display_result("正进行文件关键词替换")
            #递归处理存储目录下的文件
            self.recurse_transfer_file(self.save_dir)
            self.display_result("全部操作完毕")
        except Exception as e:
            self.display_result("==================出现异常，请钉钉联系管理员==================")
            self.display_result(str(e))
            self.display_result("=============================================================")



if __name__ == "__main__":
    app = QApplication(sys.argv)
    # 初始化
    myWin = MyMainForm()
    # 将窗口控件显示在屏幕上
    myWin.show()
    # 程序运行，sys.exit方法确保程序完整退出。
    sys.exit(app.exec_())